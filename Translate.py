import os
import sys

import requests
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from utils.AndroidFunc import AndroidFunc
from utils.TranslateThread import TranslateThread
from openpyxl import Workbook


# noinspection PyAttributeOutsideInit
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Translate Window')
        self.setGeometry(100, 100, 1400, 300)
        self.setWindowOpacity(0.9)
        self.country = ['en', 'zh-CN', 'zh-TW', 'ko', 'de', 'ja', 'ru', 'es', 'it', 'fr', 'vi', 'pt', 'fa']
        self.all_data = []
        self.translate_info = []
        self.status = self.statusBar()
        # self.menu_bar = self.menuBar()
        self.status.showMessage(f'{AndroidFunc.get_day_soup()} ———— 欢迎使用Stroke Tool,祝您使用愉快..', 10000)
        self.gui_init()
        self.on_clicked_listen()

    def gui_init(self):
        self.my_window = QWidget()
        self.setCentralWidget(self.my_window)
        # 创建一个垂直布局
        self.vbox = QVBoxLayout(self.centralWidget())
        # 创建多个水平布局
        self.hbox_1 = QHBoxLayout()
        self.hbox_2 = QHBoxLayout()
        # self.hbox_3 = QHBoxLayout()
        # self.hbox_4 = QHBoxLayout()
        self.vbox.addLayout(self.hbox_1)
        self.vbox.addLayout(self.hbox_2)
        # self.vbox.addLayout(self.hbox_3)
        # self.vbox.addLayout(self.hbox_4)

        self.need_translate_entry = QLineEdit(self)
        self.need_translate_entry.setFixedSize(1200, 20)
        self.translate_btn = QPushButton('翻译', self)
        self.choose_btn = QPushButton('选择待翻译文件', self)
        self.export_btn = QPushButton('导出Excel', self)
        self.hbox_1.addWidget(self.need_translate_entry)
        self.hbox_1.addWidget(self.choose_btn)
        self.hbox_1.addWidget(self.translate_btn)
        self.hbox_1.addWidget(self.export_btn)

        # table显示区
        self.table_log = QTableWidget()
        self.table_log.setColumnCount(len(self.country))
        self.table_log.setHorizontalHeaderLabels(
            self.country)
        # self.table_log.setColumnWidth(0, 150)
        # self.table_log.setColumnWidth(1, 70)
        # self.table_log.setColumnWidth(2, 70)
        # self.table_log.setColumnWidth(3, 70)
        self.table_log.setColumnWidth(4, 150)
        self.hbox_2.addWidget(self.table_log)

        self.setLayout(self.vbox)

        self.show()

    def on_clicked_listen(self):
        self.translate_btn.clicked.connect(self.do_translate)
        self.choose_btn.clicked.connect(self.get_config_file_path)
        self.export_btn.clicked.connect(self.get_table_log_data)

    def get_key_data(self, path):
        with open(path, 'r', encoding='utf-8') as fp:
            data = [line.strip() for line in fp]
        return data

    def get_config_file_path(self):
        path = AndroidFunc.get_file_path()
        if path[0]:
            self.need_translate_entry.setText(f'address={path[0]}')
        else:
            pass

    def do_translate(self):
        country_list = self.country
        need_key = self.need_translate_entry.text()
        if need_key:
            self.thread_start([self.country, need_key])

    def thread_start(self, key):
        self.install_adb = TranslateThread(key)
        self.install_adb.output.connect(self.on_output_received)
        self.install_adb.status_output.connect(self.on_status_output_received)
        self.install_adb.start()

    def on_output_received(self, translate_key, output):
        row = self.table_log.rowCount()
        self.table_log.insertRow(row)
        for col in range(len(output)):
            item = QTableWidgetItem(output[col])
            item.setTextAlignment(Qt.AlignCenter)
            self.table_log.setItem(row, col, item)
        self.table_log.setCurrentCell(row, 0)
        self.table_log.scrollToBottom()
        output.insert(0, translate_key)
        self.all_data.append(output)

    def on_status_output_received(self, translate_key, schedule):
        if schedule + 1 < len(self.country):
            self.status.showMessage(
                f'当前正在翻译:_>{translate_key}<_,当前翻译进度为{schedule + 1}/{len(self.country)},请等待翻译结束后查看结果~',
                1000)
        else:
            self.status.showMessage(f'当前_>{translate_key}<_已经翻译完成，请查看结果~', 10000)

    def get_table_log_data(self):
        print(self.all_data)
        try:
            os.remove(f'{AndroidFunc.get_desktop()}/Export_Translate.xlsx')
        except BaseException as e:
            print(e)
        finally:
            wb = Workbook()
            # 激活第一个工作表
            ws = wb.active
            self.country.insert(0, 'need_translate_word')
            for idx, value in enumerate(self.country, start=1):
                ws.cell(row=1, column=idx, value=value)
            for i in range(0, len(self.all_data)):
                for idx, value in enumerate(self.all_data[i], start=1):
                    ws.cell(row=i + 2, column=idx, value=value)
            wb.save(f'{AndroidFunc.get_desktop()}/Export_Translate.xlsx')
            del self.country[0]
            self.status.showMessage('导出成功,已导出到桌面名为Export_Translate.xlsx,请到桌面查看~')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    sys.exit(app.exec_())
